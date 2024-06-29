import pytest
from models import Agreements
from datetime import datetime
from graph_credit import calculating_payments
import openpyxl
 
eps = 0.001

def test_1():
    agreements = Agreements()
    agreements.principal = 1_000_000
    agreements.term = 24
    agreements.interest = 12
    agreements.datetime_activation = datetime.now()
    res = calculating_payments(agreements)
    workbook = openpyxl.load_workbook('./files/1.xlsx')
    sheet = workbook.active
 
    for row in list(sheet.iter_rows(values_only=True))[13:13+agreements.term]:
        month_ind = int(row[7] - 1)
        princip_payment = row[8]
        interest_payment = row[9]
        assert princip_payment + interest_payment == row[10]
        assert res[month_ind]['principal_payment'] - row[8] <= eps
        assert res[month_ind]['interest_payment'] - row[9] <= eps

def test_2():
    agreements = Agreements()
    agreements.principal = 10_000
    agreements.term = 8
    agreements.interest = 90
    agreements.datetime_activation = datetime.now()
    res = calculating_payments(agreements)
    workbook = openpyxl.load_workbook('./files/2.xlsx')
    sheet = workbook.active
 
    for row in list(sheet.iter_rows(values_only=True))[13:13+agreements.term]:
        month_ind = int(row[7] - 1)
        princip_payment = row[8]
        interest_payment = row[9]
        assert princip_payment + interest_payment == row[10]
        assert res[month_ind]['principal_payment'] - row[8] <= eps
        assert res[month_ind]['interest_payment'] - row[9] <= eps

def test_3():
    agreements = Agreements()
    agreements.principal = 500_000
    agreements.term = 36
    agreements.interest = 22
    agreements.datetime_activation = datetime.now()
    res = calculating_payments(agreements)
    workbook = openpyxl.load_workbook('./files/3.xlsx')
    sheet = workbook.active
 
    for row in list(sheet.iter_rows(values_only=True))[13:13+agreements.term]:
        month_ind = int(row[7] - 1)
        princip_payment = row[8]
        interest_payment = row[9]
        assert princip_payment + interest_payment == row[10]
        assert res[month_ind]['principal_payment'] - row[8] <= eps
        assert res[month_ind]['interest_payment'] - row[9] <= eps